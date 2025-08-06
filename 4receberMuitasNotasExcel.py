from lxml import etree
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
import os
import locale

# Configurações
diretorio_xml = "./xmls"
arquivo_excel = "notas_fiscais.xlsx"
ns = {'ns': 'http://www.portalfiscal.inf.br/nfe'}
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

def processar_xml(arquivo_xml):
    """Processa um único arquivo XML e retorna os dados formatados"""
    try:
        tree = etree.parse(arquivo_xml)
        root = tree.getroot()
        ns = {'ns': 'http://www.portalfiscal.inf.br/nfe'}
        
        return {
            'emitente': root.xpath('//ns:emit/ns:xNome/text()', namespaces=ns)[0],
            'vencimento': datetime.strptime(
                root.xpath('//ns:cobr/ns:dup/ns:dVenc/text()', namespaces=ns)[0], 
                "%Y-%m-%d"
            ).strftime("%d/%m/%Y"),
            'valor': f"R$ {float(root.xpath('//ns:total/ns:ICMSTot/ns:vNF/text()', namespaces=ns)[0]):.2f}".replace('.', ',')
        }
    except Exception as e:
        print(f"Erro ao processar {arquivo_xml}: {str(e)}")
        return None

# Criar Excel
wb = Workbook()
ws = wb.active
ws.title = "NF-e"

# Cabeçalhos
cabecalhos = ["Emitente", "Vencimento", "Valor"]
ws.append(cabecalhos)

# Formatar cabeçalhos
for col in range(1, 5):
    ws.cell(row=1, column=col).font = Font(bold=True)

# Processar XMLs
for idx, arquivo in enumerate([f for f in os.listdir(diretorio_xml) if f.lower().endswith('.xml')], start=1):
    dados = processar_xml(os.path.join(diretorio_xml, arquivo))
    if dados:
        ws.append([
            dados['emitente'],
            dados['vencimento'],
            dados['valor']
        ])

# Ajuste automático das colunas
for col in ws.columns:
    max_length = max(len(str(cell.value)) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_length + 2

# Salvar
wb.save(arquivo_excel)
print(f"✔ Planilha gerada: {arquivo_excel}\nTotal de notas: {ws.max_row - 1}")