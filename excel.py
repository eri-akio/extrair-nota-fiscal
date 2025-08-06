from lxml import etree
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

# Carregar o arquivo XML
xml_file = "nota.xml"
tree = etree.parse(xml_file)
root = tree.getroot()

# Definir o namespace da NF-e
ns = {'ns': 'http://www.portalfiscal.inf.br/nfe'}

# Extrair os dados
nomeEmitente = root.xpath('//ns:emit/ns:xNome/text()', namespaces=ns)[0]
vNF = root.xpath('//ns:total/ns:ICMSTot/ns:vNF/text()', namespaces=ns)[0]
data_venc = root.xpath('//ns:cobr/ns:dup/ns:dVenc/text()', namespaces=ns)[0]
data_br = datetime.strptime(data_venc, "%Y-%m-%d").strftime("%d/%m/%Y")

# Criar um arquivo Excel
wb = Workbook()
ws = wb.active
ws.title = "Dados NF-e"

# Definir cabeçalhos
cabecalhos = ["Nome", "Valor", "Data"]

# Escrever cabeçalhos (em negrito)
for col_num, cabecalho in enumerate(cabecalhos, start=1):
    ws.cell(row=1, column=col_num, value=cabecalho).font = Font(bold=True)

# Escrever dados em uma única linha (linha 2)
ws.cell(row=2, column=1, value=nomeEmitente)
ws.cell(row=2, column=2, value=f"R$ {float(vNF):.2f}")
ws.cell(row=2, column=3, value=data_br)

# Ajustar largura das colunas
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15

# Salvar o arquivo
nome_arquivo = "dados_nfe.xlsx"
wb.save(nome_arquivo)

print(f"Planilha gerada com sucesso: {nome_arquivo}")